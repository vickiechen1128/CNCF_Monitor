# -*- coding: utf-8 -*-
"""生成 MetricCenter 周进度跟踪表（MVP / v0.2 双版本维度），报告周期 2026-08-28 ~ 2026-09-03。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

ACCENT = "1F4E79"        # 深蓝主色
ACCENT_LIGHT = "D9E2F3"  # 浅蓝
GREY = "F2F2F2"
MVP_FILL = PatternFill("solid", fgColor="E2EFDA")    # MVP 绿
V02_FILL = PatternFill("solid", fgColor="FCE4D6")    # v0.2 橙
HDR_FILL = PatternFill("solid", fgColor=ACCENT)
HDR_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color=ACCENT)
SUB_FONT = Font(name="微软雅黑", size=10, color="595959")
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = Workbook()

def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 24

def put_rows(ws, start_row, rows, widths, ver_col=None):
    r = start_row
    for row in rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if ver_col and i == ver_col:
                v = str(val)
                if "v0.2" in v and "MVP" in v:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif "v0.2" in v:
                    cell.fill = V02_FILL
                elif "MVP" in v:
                    cell.fill = MVP_FILL
        r += 1
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sheet_title(ws, title, subtitle, ncols):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    ws.row_dimensions[2].height = 18

# ============ Sheet 1 周报总览 ============
ws = wb.active
ws.title = "周报总览"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 70

ws.merge_cells("B2:D2")
ws["B2"] = "MetricCenter 项目周进度跟踪"
ws["B2"].font = Font(name="微软雅黑", size=18, bold=True, color=ACCENT)
ws.row_dimensions[2].height = 34
ws.merge_cells("B3:D3")
ws["B3"] = "报告周期：2026-08-28 ~ 2026-09-03 ｜ 编制日期：2026-09-03 ｜ 数据来源：Git 提交记录、各模块 PRD Change Log、执行记录（docs/05-execution-records）"
ws["B3"].font = SUB_FONT

r = 5
ws.cell(row=r, column=2, value="一、本周关键结论").font = Font(name="微软雅黑", size=12, bold=True, color=ACCENT)
conclusions = [
    "1. v0.2 版本范围正式定版（09-02 与产品负责人逐项确认，D1~D13）：M03/M10 整体后移 v0.3；filter 自动纳入（决策53）、Job 多网域扇出（决策54）、网域覆盖表保留 v0.2；克隆 Job 移出、草稿批量提交/业务健康度看板/service_discovery 挪 v0.3；新增实例级端口覆盖 scrape_port 与 cAdvisor 容器监控口径；M07 IP 推导 / M06 ip_cidrs 后移 v0.3。",
    "2. MVP 告警分发最小闭环补齐（决策59/60）：alertmanager.yml 文件挂载 + 静默极简 UI，并纳入 M09 变更确认流水线；M08 已完成 L2/L3 规划派生，进入待开发状态。",
    "3. M06 轻量认证 + 租户/用户管理（Track B）完成开发并合并 develop（PR #48）：登录/Token/中间件、用户 CRUD、登录日志、前端登录页与路由守卫，安全整改残项（H-1/H-2/M-1/M-2）同步闭环。",
    "4. 采集状态回显链路打通口径（决策47）：安装确认不再阻断 target 生成；M01 Job 回显 + M07 三态 badge + M02 coverage API 三模块口径已对齐（09-02 A 方案），原型已同步。",
]
r += 1
for c in conclusions:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=2, value=c)
    cell.font = BODY_FONT
    cell.alignment = WRAP
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
ws.cell(row=r, column=2, value="二、本周工作量统计（公式实时统计自各明细 Sheet）").font = Font(name="微软雅黑", size=12, bold=True, color=ACCENT)
r += 1
stats = [
    ("PRD 需求变更条目数", "=COUNTA(需求场景与PRD变更!A5:A30)", "条（明细见「需求场景与PRD变更」）"),
    ("其中：MVP 相关", "=COUNTIF(需求场景与PRD变更!C5:C30,\"*MVP*\")", "条"),
    ("其中：v0.2 相关", "=COUNTIF(需求场景与PRD变更!C5:C30,\"*v0.2*\")", "条（含 v0.2 范围定版）"),
    ("原型设计更新条目数", "=COUNTA(原型设计进展!A5:A20)", "条（明细见「原型设计进展」）"),
    ("开发跟踪条目数", "=COUNTA(开发进度跟踪!A5:A30)", "条（明细见「开发进度跟踪」）"),
    ("已合并 develop", "=COUNTIF(开发进度跟踪!F5:F30,\"已合并 develop\")", "条"),
    ("本周关键产品决策", "=7+6", "项（决策44/47/48/51/52/53/54/55/56/57/59/60 + v0.2 定版 D1~D13）"),
]
ws.cell(row=r, column=2, value="指标").font = HDR_FONT
ws.cell(row=r, column=2).fill = HDR_FILL
ws.cell(row=r, column=2).border = BORDER
ws.cell(row=r, column=2).alignment = CENTER
ws.cell(row=r, column=3, value="数值").font = HDR_FONT
ws.cell(row=r, column=3).fill = HDR_FILL
ws.cell(row=r, column=3).border = BORDER
ws.cell(row=r, column=3).alignment = CENTER
ws.cell(row=r, column=4, value="说明").font = HDR_FONT
ws.cell(row=r, column=4).fill = HDR_FILL
ws.cell(row=r, column=4).border = BORDER
ws.cell(row=r, column=4).alignment = CENTER
r += 1
for name, formula, note in stats:
    ws.cell(row=r, column=2, value=name).font = BODY_FONT
    ws.cell(row=r, column=2).border = BORDER
    fcell = ws.cell(row=r, column=3, value=formula)
    fcell.font = BOLD_FONT
    fcell.border = BORDER
    fcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=4, value=note).font = SUB_FONT
    ws.cell(row=r, column=4).border = BORDER
    r += 1

r += 1
ws.cell(row=r, column=2, value="三、Sheet 索引与周度更新方法").font = Font(name="微软雅黑", size=12, bold=True, color=ACCENT)
r += 1
guide = [
    ("需求场景与PRD变更", "按模块 × 产品版本（MVP/v0.2）列出本周 PRD 变更、对应用户需求场景（角色/用户故事）与关联决策"),
    ("原型设计进展", "本周原型新增/变更功能点与 PRD 对齐状态、待对齐项"),
    ("开发进度跟踪", "开发项状态跟踪：每周例会前更新「状态」下拉与「备注」，新增条目直接在空行追加"),
    ("版本功能矩阵", "MVP 与 v0.2 功能范围定版快照（2026-09-02 口径），用于与业务需求方对账"),
]
for name, desc in guide:
    ws.cell(row=r, column=2, value=name).font = BOLD_FONT
    ws.cell(row=r, column=2).border = BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=3, value=desc).font = BODY_FONT
    ws.cell(row=r, column=3).alignment = WRAP
    ws.cell(row=r, column=3).border = BORDER
    ws.cell(row=r, column=4).border = BORDER
    r += 1

# ============ Sheet 2 需求场景与PRD变更 ============
ws2 = wb.create_sheet("需求场景与PRD变更")
sheet_title(ws2, "需求场景与 PRD 变更明细", "报告周期 2026-08-28 ~ 2026-09-03 ｜ 来源：各模块 PRD Change Log（docs/02-product-requirements/Modules/）", 10)
hdr = ["序号", "模块", "产品版本", "PRD 版本", "变更日期", "用户需求场景（角色 / 故事）", "本周变更内容", "关联决策", "PRD 状态", "对开发 / 业务方影响"]
for i, h in enumerate(hdr, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, len(hdr))

rows2 = [
    [1, "M03 网关与认证", "MVP", "v1.2 → v1.3", "08-28 / 09-02",
     "运维工程师登录平台需身份认证（访问控制门），MVP 不再是裸奔无鉴权",
     "决策44落版：MVP 轻量认证（用户名+密码登录、会话 Token、认证中间件、初始 admin 种子、前端登录页+路由守卫），Track B+ 免高保真原型；v1.3 将统一入口/Ingestion 路由后移 v0.3（Edge Agent 直连中心，v0.2 不依赖网关），MVP 认证范围不变",
     "决策44", "dev-ready", "本周已完成开发并合并 develop（见开发进度跟踪）"],
    [2, "M06 系统与平台管理", "MVP", "v2.3 → v2.4", "08-28 / 08-29",
     "平台管理员管理租户/用户账号、重置密码、查看登录日志；仅 admin/user 两级访问控制",
     "决策44落版：MVP 纳入轻量认证+租户/用户管理子集（User/LoginLog 数据模型、用户 CRUD+重置密码+软删、登录日志接口）；v2.4 收割 dev-feedback FB-04~07：H-1 生产模式拒绝默认密码、M-1 登录限流、M-2 防账号枚举、侧边栏导航命名定版（租户管理/网域管理/用户管理/登录日志）",
     "决策44；FB-04~07", "dev-ready", "已开发并合并（PR #48），可安排验收"],
    [3, "M01 监控策略与指标管理", "MVP", "v3.27", "08-28",
     "运维配完采集 Job 后想即时知道实例是否采到数据（M01-OPS-08）；不再要求先登记 Exporter 安装才能下发（M01-OPS-04 改写）",
     "决策47落版（采集状态回显前置）：①安装确认降级为可选登记，unconfirmed 不再阻断 target 生成；②Job 实例采集状态回显——状态列（待采集/up/down/unknown）+在线数汇总+down 提醒，数据经 M02 /api/v1/targets 代理只读获取；③M07 badge 三态化口径同步",
     "决策47-1/2/3", "ready", "原型已对齐 v3.28；L2/L3 规划已派生，待排期开发"],
    [4, "M07 监控对象管理", "MVP", "v2.23", "08-28",
     "运维/业务希望在线维护业务分组字典，而不是改 business_domains.yaml 重启服务（M07-OPS-10 新增）",
     "决策48落版：业务管理页提级 MVP（列表+登记+受限编辑+停用），新增 BusinessDomain 实体与 POST/PUT 接口；红线硬化——biz_code 永不可改、停用不删除、infra 禁止停用删除；yaml 降级为首次启动 seed",
     "决策48", "ready", "原型业务管理页已补（08-31），待开发"],
    [5, "M08 告警收敛与通知管理", "MVP", "v1.5 → v1.7", "08-31",
     "规则下发后告警要真正能发到接收人（端到端闭环）；值班时对告警快速静默",
     "决策59：MVP 告警分发最小闭环——alertmanager.yml 文件挂载（整文件上传/粘贴+amtool check-config 校验+版本留痕，校验失败不落库不 reload）+静默管理极简 UI（API 直调）；决策60：alertmanager.yml 作为管理域 scope 产物纳入 M09 变更确认流水线、不扇出；决策55/56：告警状态页归 M08、通知状态代理服务端授权集合过滤、静默 matcher 授权校验防越权",
     "决策55/56/59/60", "ready", "已派生 L2/L3（09-02），feat/module-08-alert-dispatch 分支待开发；原型待按 v1.7 升级"],
    [6, "M09 网域与边缘配置中心", "MVP", "v1.52", "08-31",
     "告警配置变更与采集配置一样走「生成→人工确认→下发→reload」，全程可追溯",
     "决策60落版：alertmanager.yml 纳入 ConfigDraft→人工确认→下发流水线，change_status 回写 M08；不参与按网域扇出（中心 Alertmanager 全局单例）；MVP 统一人工确认，低风险自动通过降级为后续版本预留",
     "决策60", "ready", "影响 M09 生成器范围，与 M08 开发联动"],
    [7, "M02 查询中心", "MVP", "v1.7 → v1.8", "08-31 / 09-02",
     "架构师在资源列表看「采集状态」badge，需要可信的口径（采集中/已下发未采到/未监控）",
     "决策55/56/57落版：注入三层语义——tenant_id 硬隔离强制注入、network_domain 授权集合收敛（授权=全部网域时不注入 matcher，跨网域聚合天然成立）、前端筛选不承担安全职责；存储可替换性决策点（中心预留替换 VictoriaMetrics）；v1.8 coverage 三态判定口径修订（用户拍板 A 方案）：不感知 M09 下发时序，选中未采到统一归「已下发未采到」",
     "决策55/56/57；A方案", "ready", "coverage API 是 M07 badge 数据源，planner 阻塞项已闭环；原型待按 v1.8 修订"],
    [8, "M07 监控对象管理", "MVP", "v2.25", "09-02",
     "同 7：三态 badge 口径在资源台账侧落地",
     "coverage 三态口径同步修订（提醒文案改「含变更未确认下发情形」）；五类默认标签模板补齐 resource_id → resource_id 稳定身份映射（coverage 聚合回连唯一稳定键，hostname 仅为可读别名）",
     "决策47-3", "ready", "原型 ResourcesPage 文案已同步"],
    [9, "M01 监控策略与指标管理", "MVP", "v3.29", "09-02",
     "同 7：coverage 与 Job 回显的职责边界",
     "§5.10 新增「与 coverage 三态的边界」：coverage/badge 不区分「待采集」，「待采集 vs 已下发未采到」细分仅由 M01 Job 回显承担（持有 change_status）；默认模板验收收紧为必须含 resource_id",
     "决策47-3", "ready", "纯契约收紧，原型行为不变"],
    [10, "Roadmap 产品路线图", "MVP / v0.2 / v0.3", "v2.0", "08-31",
     "全局：功能-版本矩阵与各模块决策保持一致",
     "决策51~54、59 落版同步：M01 v0.2 补 filter 与多网域扇出；M09 v0.2 补扇出与 filter 实时求值；M05 v0.3 补监控大屏（Grafana 嵌入）与首页轻量图表；M08 MVP 列补告警分发最小闭环；§4.6 可视化路线改为 Grafana iframe + ECharts/AntV",
     "决策51~54/59", "—", "跨模块统一视图，作为对账基准"],
    [11, "Roadmap 产品路线图", "v0.2", "v2.2", "09-02",
     "全体：v0.2 版本范围最终定版（与产品负责人逐项确认）",
     "D1~D13：①M03/M10 整体后移 v0.3（外联需求优先级低）；②filter(53)/Job 多网域扇出(54)/网域覆盖表保留 v0.2（M09 是 v0.2 核心）；③克隆 Job 移出、草稿批量提交/业务健康度看板/service_discovery 挪 v0.3；④新增实例级端口覆盖 scrape_port 与 cAdvisor 容器监控口径（docker_sd 降级 v0.3+ 预留）；⑤M07 IP 推导/M06 ip_cidrs 后移 v0.3；⑥K8s 划域原则（overlay CNI 独立建域、VPC 原生 CNI 可并入）",
     "D1~D13", "—", "v0.2 范围冻结，可启动估算与排期"],
    [12, "M01 监控策略与指标管理", "v0.2", "v3.30", "09-02",
     "一次定义 Job 多网域生效；新增资源自动纳入采集无需编辑 Job；同机多实例不同端口（场景多见）；容器资源监控",
     "v0.2 范围收敛：克隆 Job 移出 v0.2（扇出已覆盖跨网域复用）、草稿/批量提交 UI 挪 v0.3、业务健康度看板挪 v0.3、service_discovery 降级 v0.3+ 预留；保留 filter(53)/扇出(54)/网域覆盖表；新增 scrape_port 实例级端口覆盖（M09 生成期解析，Job 级端口映射表明确不做）；v0.2 容器监控走 cAdvisor（平台不感知容器个体）",
     "决策53/54；D3~D7", "ready", "v0.2 核心模块，待估算排期"],
    [13, "M09 网域与边缘配置中心", "v0.2", "v1.53", "09-02",
     "配置生成侧：端口自动解析、K8s 集群如何纳管",
     "target 端口解析链补 scrape_port 优先级（Resource.scrape_port → 网域覆盖表 → 映射默认 → 模板默认）；新增 K8s 划域备忘（overlay CNI 集群独立建域、vmagent Deployment/DaemonSet 复用 agent_pull 机制零改动；VPC 原生 CNI 可并入 VM 网域）；扇出/filter 维持 v0.2",
     "D5/D11", "ready", "v0.2 核心模块，待估算排期"],
    [14, "M07 监控对象管理", "v0.2", "v2.26", "09-02",
     "v0.2 导入/录入资源仍显式指定网域（与 MVP 一致，最干净）；实例端口跟随资源走",
     "网域归属 IP 推导（决策52/58，含未决队列规则化与未匹配 IP 对账视图）由 v0.2 后移 v0.3；新增可选 scrape_port 字段（登记/编辑/Excel 导入可选列，留空走 M09 端口解析链）",
     "D5/D9", "ready", "导入模板需加可选 scrape_port 列"],
    [15, "M06 系统与平台管理", "v0.2", "v2.6", "09-02",
     "多 VPC/混合云/K8s 场景下「怎么划网域」需要指导原则",
     "ip_cidrs 及 IP 推导配套后移 v0.3；新增划域指导原则——以可达性同质性+故障自治单元划域，禁止按业务/团队建域；K8s 按 CNI 选型（overlay 独立建域 zone_type 增 k8s，VPC 原生并入 VM 网域）；租户数据模型/租户-网域关联保留 v0.2",
     "D9/D11", "dev-ready", "划域原则可直接用于客户部署咨询"],
    [16, "M03 网关与认证", "v0.2 → v0.3", "v1.3", "09-02",
     "Edge Agent 直连中心拉配置，v0.2 不需要网关层",
     "统一入口路由/Ingestion 路由由 v0.2 后移 v0.3（随 M10 一并后移）",
     "D1", "dev-ready", "v0.2 无本模块开发任务"],
    [17, "M10 监控源登记册", "v0.2 → v0.3", "v1.2", "09-02",
     "外部 Prometheus 借道汇聚需求后移",
     "监控源 CRUD/外部 Remote Write/Ingestion Gateway/标签注入整体由 v0.2 后移 v0.3（外联/异构接入优先级低）；v0.4（Zabbix/云监控 Adapter）不变",
     "D1", "设计中", "v0.2 无本模块开发任务"],
    [18, "M01 / M09", "v0.2", "v3.28 / v1.51", "08-31",
     "跨网域复用不再手工克隆 Job；M07 新增资源匹配即自动纳入 targets",
     "决策53：filter 实例属性筛选模式由 v0.3 提前至 v0.2（每生成周期实时求值）；决策54：Job 网域绑定放宽为网域集合，M09 按域自动拆分 scrape_configs/targets/变更单，MVP 存量单值自动迁移",
     "决策53/54", "ready", "已在 09-02 v0.2 定版中确认保留"],
]
put_rows(ws2, 5, rows2, [6, 22, 13, 13, 13, 38, 60, 13, 11, 26], ver_col=3)
ws2.freeze_panes = "A5"

# ============ Sheet 3 原型设计进展 ============
ws3 = wb.create_sheet("原型设计进展")
sheet_title(ws3, "原型设计新增 / 变更功能点", "报告周期 2026-08-28 ~ 2026-09-03 ｜ 来源：design/module-mvp-demo 分支提交与 docs/prototypes/", 8)
hdr3 = ["序号", "模块", "产品版本", "原型版本", "更新日期", "本周新增 / 变更功能点", "与 PRD 对齐状态", "待办 / 下一步"]
for i, h in enumerate(hdr3, 1):
    ws3.cell(row=4, column=i, value=h)
style_header(ws3, 4, len(hdr3))
rows3 = [
    [1, "M01 监控策略与指标管理", "MVP", "v3.28", "08-31",
     "ScrapeJobs 页新增「采集状态」列（在线数/待采集/up/down 汇总 + down 提醒文案）；安装确认由强制闸门改为「可选登记」交互（决策47-1/47-2）",
     "已对齐 v3.28", "09-01 已完成原型结构债清理（ReviewNote 标题去 PRD/决策引用、Alert 收敛）；v3.29/v3.30 为契约收紧，原型行为不变"],
    [2, "M02 查询中心", "MVP", "v1.5", "08-31",
     "查询页 / 目标页原型：targets、coverage、envelope、注入骨架四类 MVP 能力",
     "原型 v1.5；PRD 已到 v1.8", "待按 v1.8 修订（注入三层语义 + coverage 三态口径）"],
    [3, "M05 自定义前端门户", "v0.3", "v1.1", "08-31",
     "可视化大屏页原型（Grafana iframe 嵌入 + 预置仪表盘模板，一级导航第 2 位）；首页概览 Dashboard + 新用户引导（开箱动线：登记网域→导入资源→建 Job→下发→查指标）（决策51）",
     "原型 v1.1；PRD v1.2", "待按 v1.2 修订导航层级与双入口"],
    [4, "M06 系统与平台管理", "MVP", "v2.5", "08-31",
     "网域登记页原型（Track B 增量）；租户/用户管理原型：登录页、用户管理（角色控制/删除）、租户查看/编辑、登录日志页",
     "已对齐 v2.5 ✅", "已随 PR #48 开发落地；v2.6 仅划域原则文字增量，原型行为不变"],
    [5, "M07 监控对象管理", "MVP", "v2.24 → v2.25", "08-31 / 09-02",
     "资源页 + 标签模板原型对齐；新增业务管理页 + 采集状态三态概览（菜单/列序，决策47-3/48）；ResourcesPage 文案同步 resource_id 稳定身份",
     "已对齐 v2.25 ✅", "待按 v2.26 补 scrape_port 可选列（登记表单/编辑/导入模板）"],
    [6, "M08 告警收敛与通知管理", "MVP", "v1.2", "08-31 / 09-01",
     "告警收敛边界修订（决策55）；对齐 v1.7 评审记录：校验失败不落库交互（决策59/60）+ 原型结构债清理",
     "原型 v1.2；PRD 已到 v1.7", "待按 v1.7 升级：alertmanager.yml 文件挂载页 + 静默极简 UI + 变更确认联动"],
    [7, "M09 网域与边缘配置中心", "MVP / v0.2", "v1.51 → v1.52", "08-31 / 09-01",
     "配置预览页原型（决策53/54 扇出/filter 口径）；对齐 v1.52：配置产物形态/边缘流程修订（决策60 alertmanager.yml 入流水线）+ 表格列治理（10→8 列）",
     "原型 v1.51；决策60 待原型对齐", "待对齐 v1.53（端口解析链展示）"],
]
put_rows(ws3, 5, rows3, [6, 22, 13, 15, 13, 56, 20, 34], ver_col=3)
ws3.freeze_panes = "A5"

# ============ Sheet 4 开发进度跟踪 ============
ws4 = wb.create_sheet("开发进度跟踪")
sheet_title(ws4, "开发进度跟踪（每周例会更新）", "报告周期 2026-08-28 ~ 2026-09-03 ｜ 状态列带下拉选项，新条目直接在空行追加", 10)
hdr4 = ["序号", "模块", "产品版本", "开发项", "分支 / PR", "状态", "日期", "关联执行记录 / 说明", "下周计划", "备注"]
for i, h in enumerate(hdr4, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, len(hdr4))
rows4 = [
    [1, "M00 工程基础设施", "工程（跨版本）", "业务代码符号地图 repo-map + 新鲜度门禁（pre-commit hook + CI 双门禁）",
     "develop", "已合并 develop", "08-28",
     "scripts/repo-map、check-repo-map.sh；Agent 排障定位入口", "—", "review-precheck 报告同步纳入新鲜度项"],
    [2, "M00 工程基础设施", "工程（跨版本）", "Gitflow 版本基线 tag 规范（baseline/vX.Y-*）与整版回退路径",
     "develop", "已合并 develop", "08-28",
     "docs/03-engineering-standards/06 §2.5/§6.6/§11", "—", "联调出口打 tag 作为整版回退锚点"],
    [3, "MVP 联调", "MVP", "integration/v0.1 联调分支验收合入 develop + repo-map 同步",
     "PR #47", "已合并 develop", "08-28",
     "integration/v0.1 分支已删除", "—", "首个版本联调闭环"],
    [4, "MVP 交付包", "MVP", "中心一体化交付包打包能力（metric-center + prometheus + UI），支持 Custom UI 独立访问；后端 API CORS 支持；Prometheus 种子配置路径修正",
     "develop", "已合并 develop", "08-28",
     "docs/06-mvp-e2e-testing/ 部署拓扑决策 A2", "—", "支撑 M09 local 通道端到端联调"],
    [5, "M06 轻量认证 + 租户/用户管理", "MVP", "后端：认证 API+中间件（RequireAdmin 路由挂接）、用户 CRUD+重置密码+登录日志查询+DELETE 软删、租户只读+编辑子集、admin 种子 upsert、User/Session/LoginLog 模型迁移；前端：登录页+路由守卫+401 拦截、用户管理页（角色控制/删除/顶部角色Tag）、租户查看/编辑页、登录日志页；sec-01 安全整改残项（H-1/H-2 seed&Role）",
     "feat/module-06-domain-registry → PR #48", "已合并 develop", "08-29",
     "docs/05-execution-records/module-06/track-b-increment-decision-44/", "业务需求方验收登录与用户管理流程", "Track B（dev-ready 直派开发）首个完整范例；FB-05/06/07 同步闭环"],
    [6, "M08 告警分发最小闭环", "MVP", "L2/L3 规划产物派生（决策59/60：alertmanager.yml 文件挂载 + 静默 UI + 纳入 M09 变更确认）+ 校验失败不落库口径闭环",
     "feat/module-08-alert-dispatch → PR #52", "规划完成，待开发", "09-02",
     "docs/05-execution-records/module-08/（review-round1、design-decisions）", "按 task-sequence 启动开发", "与 M09 变更确认流水线联动"],
    [7, "M01/M07 采集状态回显", "MVP", "决策47 阻断闭环 + L2/L3 派生：LabelTemplate 补 resource_id、覆盖率三态口径落版",
     "PR #52（文档）", "规划完成，待开发", "09-02",
     "docs/05-execution-records/module-01/dev-feedback.md", "排期开发 Job 回显 + badge 三态", "原型已就绪（M01 v3.28 / M07 v2.25）"],
    [8, "M01 监控策略（存量）", "MVP", "feat/module-01-strategy 分支本周无新提交（ScrapeJob/拨测等已在此前迭代交付）",
     "feat/module-01-strategy", "暂停（待新需求）", "—",
     "—", "承接决策47 回显需求后重启", "—"],
    [9, "M07 / M09（存量 feat 分支）", "MVP", "feat/module-07-resource-management、feat/module-09-config-center 本周无新提交",
     "各 feat 分支", "暂停（待新需求）", "—",
     "—", "M07 承接 scrape_port 列；M09 承接决策60 生成器改造", "—"],
    [10, "v0.2 边云主链路", "v0.2", "Edge Sync Agent / Job 网域扇出 / filter 实时求值 / scrape_port 端口解析链 / cAdvisor 容器监控 / 租户-网域关联",
     "未创建", "未启动", "—",
     "v0.2 范围 09-02 刚定版（D1~D13）", "工作量估算与排期；创建 feat 分支", "platform/edge-sync-agent/ 已有独立 module 骨架"],
]
put_rows(ws4, 5, rows4, [6, 22, 13, 52, 26, 15, 10, 30, 24, 26], ver_col=3)
ws4.freeze_panes = "A5"

# 状态下拉 + 空模板行
dv = DataValidation(type="list",
                    formula1='"未启动,规划中,规划完成，待开发,开发中,待评审,已合并 develop,已完成,阻塞,暂停（待新需求）"',
                    allow_blank=True, showDropDown=False)
ws4.add_data_validation(dv)
dv.add("F5:F40")
for r in range(5, 5 + len(rows4)):
    ws4.cell(row=r, column=6).alignment = CENTER
# 条件格式
green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
red = PatternFill("solid", fgColor="FFC7CE")
grey = PatternFill("solid", fgColor="E7E6E6")
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"已合并 develop"'], fill=green))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"已完成"'], fill=green))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"开发中"'], fill=yellow))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"规划完成，待开发"'], fill=yellow))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"阻塞"'], fill=red))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"暂停（待新需求）"'], fill=grey))
ws4.conditional_formatting.add("F5:F40", CellIsRule(operator="equal", formula=['"未启动"'], fill=grey))

# ============ Sheet 5 版本功能矩阵 ============
ws5 = wb.create_sheet("版本功能矩阵")
sheet_title(ws5, "MVP / v0.2 功能范围定版快照", "口径：Roadmap v2.2 §1.5 功能-版本矩阵（2026-09-02 v0.2 定版）｜ 用于与业务需求方对账", 4)
hdr5 = ["模块", "MVP 范围（定版）", "v0.2 范围（2026-09-02 定版）", "本周范围变化"]
for i, h in enumerate(hdr5, 1):
    ws5.cell(row=4, column=i, value=h)
style_header(ws5, 4, len(hdr5))
rows5 = [
    ["M07 监控对象管理",
     "五类资源 CRUD；Excel 导入（upsert + network_domain_id 必填）；状态映射（offline 排除）；标签模板；ResourceLabel 体系；三态采集 badge（本周新增口径）；业务类型归属",
     "资源可选采集端口字段 scrape_port（实例级端口覆盖，登记/编辑/Excel 导入可选列）",
     "IP 推导后移 v0.3；新增 scrape_port（D5/D9）"],
    ["M01 监控策略与指标管理",
     "CI↔默认采集器绑定；ScrapeJob（实例选择+冻结网域校验）；采集认证/TLS 最小集；Blackbox 拨测；静态指标库；application_http 业务指标采集；业务指标库；规则文件挂载（rules.yml 经 M09 下发）",
     "网域级覆盖表 CITypeExporterMappingOverride；filter 实例属性筛选（决策53，新增资源自动纳入）；Job 多网域绑定+按域扇出（决策54）；实例级端口覆盖；cAdvisor 容器资源监控",
     "克隆 Job 移出待评估；草稿批量提交/业务健康度看板/service_discovery 挪 v0.3；新增 scrape_port 与 cAdvisor 口径（D3~D7）"],
    ["M09 网域与边缘配置中心",
     "默认网域 default；单/多网域模式切换；配置生成/预览/Diff/下发；external_labels 注入；change_status 回写；规则文件生成 rules.yml；alertmanager.yml 纳入变更确认（本周新增）",
     "网域生命周期与 Token；Edge Sync Agent；按网域配置拉取；Agent 状态列表；Remote Write 参数；Job 网域扇出生成；filter 实时求值；target 端口解析链",
     "决策60 alertmanager.yml 入流水线；端口解析链补 scrape_port；K8s 划域备忘（D5/D11）"],
    ["M02 查询中心",
     "PromQL 查询代理；目标状态展示；响应 envelope；采集健康度/coverage API（供 M07 badge，本周口径修订）",
     "租户/网域上下文注入（三层语义，本周落版）",
     "coverage 三态判定口径修订（A 方案）；注入三层语义（决策55/56/57）"],
    ["M08 告警收敛与通知管理",
     "告警分发最小闭环：alertmanager.yml 文件挂载（接收人/路由/抑制低频配置）+ 静默管理极简 UI；变更经 M09 确认下发（本周新增）",
     "—（MVP 最小闭环；接收人/路由表单化 UI、告警状态页、抑制引擎在 v0.3）",
     "决策59/60 落版：MVP 由「-」变为有交付"],
    ["M06 系统与平台管理",
     "单租户默认模式 + 网域登记管理 + 轻量认证与用户管理子集（本周新增并已开发合并）",
     "租户数据模型；租户-网域关联；Tenant.multi_site_enabled",
     "决策44 轻量认证落版并开发完成；ip_cidrs 后移 v0.3；新增划域指导原则（D9/D11）"],
    ["M10 监控源登记册", "—",
     "—（监控源 CRUD/外部 Remote Write/Ingestion Gateway/标签注入整体后移 v0.3）",
     "整体后移 v0.3（D1）"],
    ["M04 自定义服务发现", "Excel Provider（由 M07 承载）", "—", "无变化"],
    ["M03 网关与认证",
     "轻量认证（登录/会话 Token/认证中间件，本周落版并开发合并）",
     "—（统一入口/Ingestion 路由后移 v0.3）",
     "决策44 落版；统一入口后移 v0.3（D1）"],
    ["M05 自定义前端门户", "—",
     "—（门户/大屏/首页在 v0.3）",
     "大屏与首页原型已先行设计（决策51），版本归属 v0.3 不变"],
]
put_rows(ws5, 5, rows5, [22, 55, 50, 34], ver_col=None)
for r in range(5, 5 + len(rows5)):
    ws5.cell(row=r, column=2).fill = MVP_FILL
    ws5.cell(row=r, column=3).fill = V02_FILL
ws5.freeze_panes = "A5"

out = "MetricCenter_周进度跟踪_2026-08-28至09-03.xlsx"
wb.save(out)
print("saved:", out)
